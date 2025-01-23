# coding: utf-8
import os
from openpyxl import Workbook, load_workbook
import logging
from typing import List, Tuple, Any, Optional

class XLSXFileManager:
    def __init__(self, file_path: str, sheet_name: str, header: List[str], unique_indices: List[int], force_delete: bool = False) -> None:
        """
        初始化 XLSXFileManager 实例。
        :param file_path: 文件名
        :param header: 第一行的表头数据
        :param unique_indices: 用于排重的数据位置数组，例如 [0, 1] 表示使用第一列和第二列的数据进行排重
        """
        self.file_path = file_path
        self.header = header
        self.unique_indices = unique_indices

        if force_delete:
            try:
                os.remove(self.file_path)
                logging.info(f"文件已删除: {self.file_path}")
            except FileNotFoundError:
                logging.error("尝试删除的文件不存在")
            except PermissionError:
                logging.error("没有权限删除文件")
            except Exception as e:
                logging.error(f"删除文件时出现错误: {e}")

        if os.path.exists(self.file_path):
            try:
                self.workbook = load_workbook(self.file_path)
            except KeyError as e:
                logging.error(f"文件加载失败，可能文件损坏或格式不正确: {e}")
            self.sheet = self.workbook[sheet_name]
        else:
            self.workbook = Workbook()
            self.sheet = self.workbook.active
            self.sheet.title = sheet_name
            self.sheet.append(self.header)
            self.workbook.save(self.file_path)

    def _is_duplicate(self, data: List[Any]) -> bool:
        """
        检查待插入的数据是否与已有数据重复。
        :param data: 待插入的数据列表
        :return: 如果数据重复返回 True，否则返回 False
        """
        for row in self.sheet.iter_rows(min_row=2, values_only=True):
            if all(row[idx] == data[idx] for idx in self.unique_indices):
                return True
        return False

    def insert_and_save(self, data: List[Any]) -> None:
        """
        插入数据并保存文件，排重后执行插入操作。
        :param data: 待插入的数据列表
        """
        if not self._is_duplicate(data):
            self.sheet.append(data)
            self.workbook.save(self.file_path)
        else:
            logging.debug(f"数据重复，未插入: {data}")

    def insert_and_save_data_list(self, data_list: List[List[Any]]) -> None:
        """
        插入多行数据并保存文件，排重后执行插入操作。
        :param data_list: 待插入的多行数据列表
        """
        for data in data_list:
            if not self._is_duplicate(data):
                self.sheet.append(data)
            else:
                logging.debug(f"数据重复，未插入: {data}")
        self.workbook.save(self.file_path)

    def read_all_rows(self) -> List[Tuple[Any, ...]]:
        """
        读取所有行的数据。
        :return: 包含所有行数据的列表
        """
        return [row for row in self.sheet.iter_rows(min_row=2, values_only=True)]

    def close(self) -> None:
        """
        关闭工作簿。
        """
        self.workbook.close()
