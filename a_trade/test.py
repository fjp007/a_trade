import os
from openpyxl import load_workbook, Workbook
import logging
import datetime
from a_trade.trade_calendar import TradeCalendar
from a_trade.db_base import Session
from a_trade.market_analysis import MarketDailyData
from a_trade.stocks_daily_data import StockDailyData
from a_trade.limit_up_data_tushare import LimitUpTushare, LimitDataSource,get_limit_info
import pandas as pd
import a_trade.settings
from typing import Optional
from a_trade.reason_concept import ReasonConceptManger
from sqlalchemy import func

total_days = 0

def find_differences_between_files(file_path1, file_path2, output_file, sheet_name, unique_indices):
    """
    比较两个 Excel 文件，找到不同的记录，并保存到一个新的文件中。
    :param file_path1: 第一个 Excel 文件路径
    :param file_path2: 第二个 Excel 文件路径
    :param output_file: 输出结果的文件路径
    :param sheet_name: 工作表名称
    :param unique_indices: 用于判断记录是否相同的列索引
    """
    def load_data(file_path):
        """直接使用 openpyxl 加载数据"""
        try:
            workbook = load_workbook(file_path, data_only=True)
            sheet = workbook[sheet_name]
            data = [tuple(cell.value for cell in row) for row in sheet.iter_rows(min_row=2)]
            workbook.close()
            return data
        except Exception as e:
            logging.error(f"读取文件 {file_path} 时出错: {e}")
            return []
        
    def process_value(value):
        # 如果是字符串且符合日期时间格式，转换为日期格式
        if isinstance(value, str):
            try:
                # 尝试将字符串解析为 datetime 对象，并只取日期部分
                return datetime.datetime.strptime(value, "%Y-%m-%d %H:%M:%S").strftime("%Y-%m-%d")
            except ValueError:
                pass  # 不是日期格式，保持原值
        return value

    def data_to_key_set(data, unique_indices):
        """
        将数据转换为基于 unique_indices 的唯一性集合，并对日期格式的数据进行格式化处理。
        如果是日期格式，形如 "2022-06-23 09:31:00"，则转化为 "2022-06-23"。
        """
        return {
            tuple(process_value(row[idx]) for idx in unique_indices) for row in data
        }

    # 读取两个文件的数据
    data1 = load_data(file_path1)
    data2 = load_data(file_path2)

    # 转换为集合，便于比较
    key_set1 = data_to_key_set(data1, unique_indices)
    key_set2 = data_to_key_set(data2, unique_indices)

    # 找到两个文件中的差异数据
    diff_keys = key_set1.difference(key_set2)  # 在两个集合中不重复的键
    print(diff_keys)

    # 从原始数据中找出对应的完整行
    diff_data = [
        row for row in data1 + data2
        if tuple(process_value(row[idx]) for idx in unique_indices) in diff_keys
    ]
    # print(diff_data)

    # 将差异数据写入新文件
    if diff_data:
        if os.path.exists(output_file):
            os.remove(output_file)  # 如果输出文件已存在，删除它
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = sheet_name
        sheet.append(["板块名称", "股票代码", "股票名称", "烂板当日成交额占比(%)", "连板数", "购买日", "购买价格", "买入当天是否封板", "买入日情绪指标", "买入日开盘(%)", "卖出日", "卖出价格", "盈利(%)"])  # 表头
        for row in diff_data:
            sheet.append(row)  # 写入每一行差异数据
        workbook.save(output_file)
        logging.info(f"差异数据已写入文件: {output_file}")
    else:
        logging.info("两个文件的记录完全相同，没有差异。")

def analysis_data(start_date, end_date):
   with Session() as session:
        # 查询符合条件的日期
        sentiment_results = session.query(
            MarketDailyData.trade_date
        ).filter(
            MarketDailyData.sentiment_index <= 40,
            MarketDailyData.trade_date.between(start_date, end_date)
        ).all()
        
        # 获取所有符合条件日期的下一个交易日
        next_trade_dates = []
        for result in sentiment_results:
            current_trade_date = result.trade_date
            next_trade_date = TradeCalendar.get_next_trade_date(current_trade_date)
            if next_trade_date:
                next_trade_dates.append(next_trade_date)
        
        # 去重处理，避免重复查询
        next_trade_dates = list(set(next_trade_dates))
        
        # 一次性查询所有下一个交易日的数据
        next_day_results = session.query(
            MarketDailyData.trade_date,
            MarketDailyData.physical_board_next_day_limit_up_rate
        ).filter(
            MarketDailyData.trade_date.in_(next_trade_dates)
        ).all()
        
        # 计算平均值
        rates = [
            result.physical_board_next_day_limit_up_rate
            for result in next_day_results
            if result.physical_board_next_day_limit_up_rate is not None
        ]
        avg_next_day_rate = sum(rates) / len(rates) if rates else None

        # 计算总体平均值
        overall_result = session.query(
            func.avg(MarketDailyData.physical_board_next_day_limit_up_rate).label('avg_overall_rate')
        ).filter(
            MarketDailyData.trade_date.between(start_date, end_date)
        ).one()
        
        avg_overall_rate = overall_result.avg_overall_rate

        # 打印结果
        print(f"Average second-day limit-up rate (sentiment_index <= 40): {avg_next_day_rate}")
        print(f"Average limit-up rate for all days between {start_date} and {end_date}: {avg_overall_rate}")

def process_excel_and_enrich_data(file_path, output_path):
    # Step 1: Read Excel data
    df = pd.read_excel(file_path)
    
    # Select relevant columns
    df = df[["板块名称", "股票代码", "股票名称", "连板数", "购买日", "购买价格", 
             "买入当天是否封板","买入日开盘(%)", "卖出日", "卖出价格", "盈利(%)"]]
    
    # Step 2: Retrieve additional data using ORM
    enriched_data = []
    session = Session()
    try:
        for _, row in df.iterrows():
            ts_code = row["股票代码"]
            purchase_date_raw = row["购买日"]
            buy_price = row["购买价格"]
            
            # Convert purchase date to "YYYYMMDD" format for query
            purchase_date = datetime.datetime.strptime(purchase_date_raw, "%Y-%m-%d %H:%M:%S").strftime("%Y%m%d")
            
            # Query the next trading day using ORM
            next_trade_date = TradeCalendar.get_next_trade_date(purchase_date)
            buy_date_stock_data = (
                session.query(StockDailyData)
                .filter(StockDailyData.trade_date == purchase_date, StockDailyData.ts_code == ts_code)
                .first()
            )
            buy_date_limit_data = (
                session.query(LimitUpTushare)
                .filter(LimitUpTushare.trade_date == purchase_date, LimitUpTushare.stock_code == ts_code)
                .first()
            )
            buy_date_pch = (buy_date_stock_data.close - buy_price)*100/buy_price

            if buy_date_limit_data:
                buy_date_first_time = buy_date_limit_data.first_time
            else:
                buy_date_first_time = None
            
            if next_trade_date:
                
                # Query the opening percentage change for the next trading day
                stock_data = (
                    session.query(StockDailyData)
                    .filter(StockDailyData.trade_date == next_trade_date, StockDailyData.ts_code == ts_code)
                    .first()
                )
                
                opening_pct_chg = (stock_data.open / stock_data.pre_close - 1)*100 if stock_data else None
                
                # Query the limit-up last time for the next trading day
                limit_up_data = (
                    session.query(LimitUpTushare)
                    .filter(LimitUpTushare.trade_date == next_trade_date, LimitUpTushare.stock_code == ts_code)
                    .first()
                )
                
                limit_up_first_time = limit_up_data.first_time if limit_up_data else None
                if limit_up_data:
                    if limit_up_data.limit_status == 'U':
                        limit_type = "涨停"
                    elif limit_up_data.limit_status == 'D':
                        limit_type = '跌停'
                    else:
                        limit_type = '炸板'
                else:
                    limit_type = '未封板'
            else:
                opening_pct_chg = None
                limit_up_first_time = None
                limit_type = None
            
            enriched_data.append({
                **row.to_dict(),
                "买入日收益": buy_date_pch,
                "买入日首次涨停时间": buy_date_first_time,
                "第二天开盘幅度(%)": opening_pct_chg,
                "第二天涨停时间": limit_up_first_time,
                "第二天涨跌停": limit_type,
                "第二天开盘价": stock_data.open,
                "第二天最高价": stock_data.high,
            })
    finally:
        session.close()
    
    # Step 3: Write enriched data to a new Excel file
    enriched_df = pd.DataFrame(enriched_data)
    enriched_df.to_excel(output_path, index=False)
    print(f"Data successfully processed and saved to {output_path}")

total_count = 0
next_limit_count = 0
total_40_count = 0
next_limit_40_count = 0


def test_rate(trade_date):
    global total_count
    global next_limit_count
    today_data_source = LimitDataSource(trade_date)
    next_trade_date = TradeCalendar.get_next_trade_date(trade_date)
    next_data_source = LimitDataSource(next_trade_date)

    for stock_info in today_data_source.limit_up_map.values():
        limit_info: Optional[LimitUpTushare] = stock_info
        stock_daily_data: Optional[StockDailyData] = today_data_source.get_daily_data(limit_info.stock_code)
        if not stock_daily_data.is_one_limit():
            if limit_info.up_stat == '1/1':
                total_count += 1
                if limit_info.stock_code in next_data_source.limit_up_map:
                    next_limit_count += 1
    

if __name__ == '__main__':
    import sys
    # from sqlalchemy import func
    # from sqlalchemy.orm import aliased
    # from a_trade.wechat_bot import WechatBot

    file2 = '/Users/fengjianpeng/Documents/Leon/Workspace/a_trade/strategy_result/Yugi_s1交易记录20220620_20241130.xlsx'
    file1 = '/Users/fengjianpeng/Documents/Yugi_s1交易记录20220620_20241130.xlsx'
    # output_file = '/strategy_result/test.xlsx'
    # # process_excel_and_enrich_data(file1, output_file)
    find_differences_between_files(file1, file2, '/strategy_result/difference.xlsx', "历史交易记录", [2, 5, 12])
    # start_date = sys.argv[1]
    # end_date = sys.argv[2]
    # analysis_data(start_date, end_date)